# -*- coding: utf-8 -*-

'''Easily import an Excel spreadsheet with headers on the top row.'''

from __future__ import (absolute_import, division, print_function,
                        unicode_literals)
from zipfile import BadZipFile
from nine import IS_PYTHON2, nimport, nine, range, str, basestring
from openpyxl import load_workbook  # easy_install -UZ openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from bag.web.exceptions import Problem
try:
    from bag.web.pyramid import _
except ImportError:
    _ = str  # and i18n is disabled.
from . import (
    get_corresponding_variable_names, raise_if_missing_required_headers)


def excel_reader(stream, worksheet_name=None, required_headers=[]):
    '''Reads an XLSX file (from ``stream``) and yields objects so you can
        access the values conveniently.

        You can pass in the ``worksheet_name`` to be read.  If not passed in or
        not present in the file, the first worksheet will be read.

        In addition, you may pass a sequence of *required_headers*, and if they
        aren't all present, KeyError is raised.

        Let's see an example. Suppose you are reading some Excel file and
        all you know is it contains the columns "E-mail", "Full Name" and
        "Gender", not necessarily in that order::

            reader = excel_reader(
                open('contacts.xlsx', mode='rb'),
                worksheet_name='Mailing',
                required_headers=['E-mail', 'Full Name', 'Gender'])
            for o in reader:
                print(o.full_name, o.e_mail, o.gender)
        '''
    try:
        wb = load_workbook(stream, data_only=True)
    except (BadZipFile, InvalidFileException) as e:
        raise Problem(
            _('That is not an XLSX file.'),
            error_title=_('Unable to read the XLSX file'), error_debug=str(e))

    # Grab either the worksheet named "Assets", or simply the first one
    if worksheet_name and worksheet_name in wb:
        sheet = wb[worksheet_name]
    else:
        sheet = wb[wb.sheetnames[0]]

    this_is_the_first_row = True
    for row in sheet.rows:
        if this_is_the_first_row:  # Read and validate the headers
            this_is_the_first_row = False
            headers = [cell.value for cell in row]
            raise_if_missing_required_headers(headers, required_headers)
            vars = get_corresponding_variable_names(headers, required_headers)
            index_of_var = {var: i for i, var in enumerate(vars)}

            class SpreadsheetRow(object):
                '''View on a spreadsheet row so you can access data as if
                    they were instance variables.
                    '''
                __slots__ = ('__cells',)

                def __init__(self, cells):
                    self.__cells = cells

                def __getattr__(self, attr):
                    content = self.__cells[index_of_var[attr]].value
                    return content

        else:
            yield SpreadsheetRow(row)
